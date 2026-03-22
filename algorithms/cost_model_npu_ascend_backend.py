from __future__ import annotations
import csv
import json
import os
import re
from functools import lru_cache
from typing import Optional, Dict, Tuple, Any, List
import logging

from config import attach_local_debug_filter

logger = logging.getLogger(__name__)
attach_local_debug_filter(logger, lambda: True)

_THIS_DIR = os.path.dirname(__file__)
_RUNTIME_DIR = os.path.join(_THIS_DIR, 'runtime_models')

# LUT defaults
_DEF_MMAD_LUT_REL       = os.path.join(_RUNTIME_DIR, 'mmad_lookup_table.json')
_DEF_SOFTMAX_LUT_REL    = os.path.join(_RUNTIME_DIR, 'softmax_lookup_table.json')
_DEF_GELU_LUT_REL       = os.path.join(_RUNTIME_DIR, 'gelu_lookup_table.json')
_DEF_LAYERNORM_LUT_REL  = os.path.join(_RUNTIME_DIR, 'layernorm_lookup_table.json')


def _candidate_mmad_lut_paths() -> list[str]:
    cand: list[str] = []
    # Environment override
    env_path = os.environ.get('TRIFORM_MMAD_LUT')
    if env_path:
        cand.append(env_path)

    # Prefer LUTs colocated with this python file (common in repo workflows)
    cand.extend([
        os.path.join(_THIS_DIR, 'mmad_lut.csv'),
        os.path.join(_THIS_DIR, 'mmad_lut.json'),
        os.path.join(_THIS_DIR, 'mmad_lookup_table.csv'),
        os.path.join(_THIS_DIR, 'mmad_lookup_table.json'),
    ])

    # Runtime-model fallbacks
    cand.extend([
        _DEF_MMAD_LUT_REL,
        os.path.join(_RUNTIME_DIR, 'mmad_lut.json'),
        os.path.join(_RUNTIME_DIR, 'mmad_lookup_table.csv'),
        os.path.join(_RUNTIME_DIR, 'mmad_lut.csv'),
    ])
    return cand


def _candidate_softmax_lut_paths() -> list[str]:
    cand: list[str] = []
    # Environment override
    env_path = os.environ.get('TRIFORM_SOFTMAX_LUT')
    if env_path:
        cand.append(env_path)

    # Prefer LUTs colocated with this python file
    cand.extend([
        os.path.join(_THIS_DIR, 'softmax_lut.csv'),
        os.path.join(_THIS_DIR, 'softmax_lut.json'),
        os.path.join(_THIS_DIR, 'softmax_lookup_table.csv'),
        os.path.join(_THIS_DIR, 'softmax_lookup_table.json'),
    ])

    # Runtime-model fallbacks
    cand.extend([
        _DEF_SOFTMAX_LUT_REL,
        os.path.join(_RUNTIME_DIR, 'softmax_lut.json'),
        os.path.join(_RUNTIME_DIR, 'softmax_lookup_table.csv'),
        os.path.join(_RUNTIME_DIR, 'softmax_lut.csv'),
    ])
    return cand


def _candidate_gelu_lut_paths() -> list[str]:
    cand: list[str] = []
    # Environment override
    env_path = os.environ.get('TRIFORM_GELU_LUT')
    if env_path:
        cand.append(env_path)

    # Prefer LUTs colocated with this python file
    cand.extend([
        os.path.join(_THIS_DIR, 'gelu_lut.csv'),
        os.path.join(_THIS_DIR, 'gelu_lut.json'),
        os.path.join(_THIS_DIR, 'gelu_lookup_table.csv'),
        os.path.join(_THIS_DIR, 'gelu_lookup_table.json'),
    ])

    # Runtime-model fallbacks
    cand.extend([
        _DEF_GELU_LUT_REL,
        os.path.join(_RUNTIME_DIR, 'gelu_lut.json'),
        os.path.join(_RUNTIME_DIR, 'gelu_lookup_table.csv'),
        os.path.join(_RUNTIME_DIR, 'gelu_lut.csv'),
    ])
    return cand


def _candidate_layernorm_lut_paths() -> list[str]:
    cand: list[str] = []
    # Environment overrides (treat RMSNorm as generic "norm" LUT as well)
    for ev in ('TRIFORM_NORM_LUT', 'TRIFORM_LAYERNORM_LUT', 'TRIFORM_RMSNORM_LUT'):
        env_path = os.environ.get(ev)
        if env_path:
            cand.append(env_path)

    # Prefer LUTs colocated with this python file
    cand.extend([
        os.path.join(_THIS_DIR, 'rmsnorm_lut.csv'),
        os.path.join(_THIS_DIR, 'rmsnorm_lut.json'),
        os.path.join(_THIS_DIR, 'layernorm_lut.csv'),
        os.path.join(_THIS_DIR, 'layernorm_lut.json'),
        os.path.join(_THIS_DIR, 'norm_lut.csv'),
        os.path.join(_THIS_DIR, 'norm_lut.json'),
        os.path.join(_THIS_DIR, 'layernorm_lookup_table.csv'),
        os.path.join(_THIS_DIR, 'layernorm_lookup_table.json'),
    ])

    # Runtime-model fallbacks
    cand.extend([
        _DEF_LAYERNORM_LUT_REL,
        os.path.join(_RUNTIME_DIR, 'rmsnorm_lut.json'),
        os.path.join(_RUNTIME_DIR, 'rmsnorm_lut.csv'),
        os.path.join(_RUNTIME_DIR, 'layernorm_lut.json'),
        os.path.join(_RUNTIME_DIR, 'layernorm_lookup_table.csv'),
        os.path.join(_RUNTIME_DIR, 'layernorm_lut.csv'),
    ])
    return cand

# --------------------------------------------------- LUT loading + interpolation --------------------------------------

def _try_read_json(path: str) -> Any:
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def _try_read_csv_rows(path: str) -> List[Dict[str, Any]]:
    with open(path, 'r', encoding='utf-8') as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=[',', '\t', ';'])
            reader = csv.DictReader(f, dialect=dialect)
            return [dict(r) for r in reader if r]
        except Exception:
            # Fallback: whitespace-separated without header -> M N K us
            f.seek(0)
            rows: List[Dict[str, Any]] = []
            for line in f:
                s = line.strip()
                if (not s) or s.startswith('#'):
                    continue
                parts = s.split()
                if len(parts) < 2:
                    continue
                rows.append({str(i): parts[i] for i in range(len(parts))})
            return rows


def _try_read_xlsx_rows(path: str) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise ImportError(f"openpyxl is required to read '{path}': {e}") from e

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = []
    for i, h in enumerate(rows[0]):
        hs = str(h).strip() if h is not None else str(i)
        header.append(hs if hs else str(i))
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if r is None:
            continue
        d: Dict[str, Any] = {}
        for h, v in zip(header, r):
            d[h] = v
        # skip fully-empty rows
        if any(v is not None and str(v).strip() != '' for v in d.values()):
            out.append(d)
    return out


def _load_table_any(path: str) -> Any:
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.json',):
        return _try_read_json(path)
    if ext in ('.xlsx', '.xlsm'):
        return _try_read_xlsx_rows(path)
    if ext in ('.csv', '.tsv', '.txt'):
        return _try_read_csv_rows(path)
    # Unknown extension: try JSON then CSV best-effort
    try:
        return _try_read_json(path)
    except Exception:
        return _try_read_csv_rows(path)


def _coerce_int(x: Any) -> Optional[int]:
    try:
        if x is None:
            return None
        if isinstance(x, bool):
            return int(x)
        if isinstance(x, (int,)):
            return int(x)
        s = str(x).strip()
        if not s:
            return None
        return int(float(s))
    except Exception:
        return None


def _coerce_float(x: Any) -> Optional[float]:
    try:
        if x is None:
            return None
        if isinstance(x, bool):
            return float(int(x))
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def _first_present(d: Dict[str, Any], keys: Tuple[str, ...]) -> Any:
    for k in keys:
        if k in d:
            return d.get(k)
        # also try lower-case
        kl = k.lower()
        for kk in d.keys():
            if str(kk).lower() == kl:
                return d.get(kk)
    return None


def _flatten_nested_numeric_dict_3d(obj: Any) -> List[Tuple[int, int, int, float]]:
    """Flatten nested dict like {M:{N:{K:us}}} into point tuples."""
    pts: List[Tuple[int, int, int, float]] = []
    if not isinstance(obj, dict):
        return pts
    for mk, mv in obj.items():
        m = _coerce_int(mk)
        if m is None or not isinstance(mv, dict):
            continue
        for nk, nv in mv.items():
            n = _coerce_int(nk)
            if n is None or not isinstance(nv, dict):
                continue
            for kk, vv in nv.items():
                k = _coerce_int(kk)
                us = _coerce_float(vv)
                if None not in (m, n, k, us):
                    pts.append((int(m), int(n), int(k), float(us)))
    return pts


def _build_index(points: List[Tuple[int, ...]]) -> Dict[Tuple[int, ...], float]:
    idx: Dict[Tuple[int, ...], float] = {}
    for p in points:
        *dims, val = p
        key = tuple(int(x) for x in dims)
        # If duplicates exist, keep the min latency (usually safer).
        try:
            v = float(val)
        except Exception:
            continue
        if key in idx:
            idx[key] = float(min(idx[key], v))
        else:
            idx[key] = float(v)
    return idx


def _mmad_distance(
    query_dims: Tuple[int, ...],
    point_dims: Tuple[int, ...],
    *,
    cube: int = 16,
    residue_weight: float = 0.05,
) -> float:
    """Distance in MMAD tile space.

    Keep MMAD interpolation closer to Ascend Cube behavior by comparing
    shapes in tile space first, then adding a small residue penalty inside a
    tile bucket.
    """
    q = tuple(int(max(1, int(x))) for x in query_dims)
    p = tuple(int(max(1, int(x))) for x in point_dims)
    cube_i = int(max(1, cube))

    q_tiles = tuple((x + cube_i - 1) // cube_i for x in q)
    p_tiles = tuple((x + cube_i - 1) // cube_i for x in p)
    q_res = tuple(x % cube_i for x in q)
    p_res = tuple(x % cube_i for x in p)

    dist = 0.0
    for qi, pi in zip(q_tiles, p_tiles):
        dist += abs(float(qi) - float(pi)) / float(max(1, qi))

    if residue_weight > 0.0:
        norm = float(max(1, cube_i - 1))
        for qi, pi in zip(q_res, p_res):
            dist += float(residue_weight) * abs(float(qi) - float(pi)) / norm

    return float(dist)


def _idw_interpolate(
    points: List[Tuple[int, ...]],
    query_dims: Tuple[int, ...],
    *,
    k_neigh: int = 8,
    power: float = 1.0,
    distance_fn: Optional[Any] = None,
) -> Optional[float]:
    """Inverse-distance weighted interpolation over sparse LUT points."""
    if not points:
        return None
    q = tuple(int(max(1, int(x))) for x in query_dims)
    # compute distances
    scored: List[Tuple[float, float]] = []
    eps = 1e-12
    for p in points:
        dims = tuple(int(x) for x in p[:-1])
        val = float(p[-1])
        if distance_fn is None:
            # relative L1 distance in raw shape space
            dist = 0.0
            for qi, pi in zip(q, dims):
                dist += abs(float(qi) - float(pi)) / float(max(1, qi))
        else:
            dist = float(distance_fn(q, dims))
        if dist <= 0.0:
            return float(val)
        scored.append((dist, val))
    scored.sort(key=lambda t: t[0])
    take = scored[: max(1, int(k_neigh))]
    num = 0.0
    den = 0.0
    for dist, val in take:
        w = 1.0 / (float(dist) ** float(power) + eps)
        num += w * float(val)
        den += w
    if den <= 0.0:
        return float(take[0][1])
    return float(num / den)


def _idw_interpolate_with_neighbors(
    points: List[Tuple[int, ...]],
    query_dims: Tuple[int, ...],
    *,
    k_neigh: int = 8,
    power: float = 1.0,
    distance_fn: Optional[Any] = None,
) -> Tuple[Optional[float], Tuple[Tuple[Tuple[int, ...], float, float], ...]]:
    if not points:
        return (None, tuple())
    q = tuple(int(max(1, int(x))) for x in query_dims)
    scored: List[Tuple[float, Tuple[int, ...], float]] = []
    eps = 1e-12
    for p in points:
        dims = tuple(int(x) for x in p[:-1])
        val = float(p[-1])
        if distance_fn is None:
            dist = 0.0
            for qi, pi in zip(q, dims):
                dist += abs(float(qi) - float(pi)) / float(max(1, qi))
        else:
            dist = float(distance_fn(q, dims))
        if dist <= 0.0:
            return (float(val), ((dims, float(val), 0.0),))
        scored.append((dist, dims, val))
    scored.sort(key=lambda t: t[0])
    take = scored[: max(1, int(k_neigh))]

    num = 0.0
    den = 0.0
    for dist, _dims, val in take:
        w = 1.0 / (float(dist) ** float(power) + eps)
        num += w * float(val)
        den += w
    if den <= 0.0:
        out = float(take[0][2])
    else:
        out = float(num / den)

    neigh = tuple((tuple(d), float(v), float(dist)) for dist, d, v in take)
    return (out, neigh)


@lru_cache(maxsize=4096)
def _warn_missing_once(tag: str, key: Tuple[int, ...], path: str) -> None:
    logger.warning(f"[{tag}-LUT] Missing key={key} in LUT='{path}'. Using LUT interpolation.")


@lru_cache(maxsize=32768)
def _log_lut_hit_once(tag: str, key: Tuple[int, ...], us: float, path: str) -> None:
    msg = f"[{tag}-LUT][HIT] key={key} us={float(us):.6f} LUT='{path}'"
    if logger.isEnabledFor(logging.INFO):
        logger.info(msg)
    else:
        logger.warning(msg)

@lru_cache(maxsize=32768)
def _log_lut_interp_once(
    tag: str,
    key: Tuple[int, ...],
    us: float,
    path: str,
    neighbors: Tuple[Tuple[Tuple[int, ...], float, float], ...],
) -> None:
    # Log the neighbor rows used so it's obvious *which* LUT rows were used to simulate this key.
    # Example neighbor element: ((M,N,K), us, dist)
    if neighbors:
        neigh_s = ', '.join(
            [f"{dims}:us={val:.6f},d={dist:.4f}" for (dims, val, dist) in neighbors]
        )
    else:
        neigh_s = ''
    msg = (
        f"[{tag}-LUT][INTERP] key={key} us={float(us):.6f} LUT='{path}'"
        + (f" neigh=[{neigh_s}]" if neigh_s else "")
    )
    if logger.isEnabledFor(logging.INFO):
        logger.info(msg)
    else:
        logger.warning(msg)


def _lut_query(
    *,
    tag: str,
    lut: Dict[str, Any],
    dims: Tuple[int, ...],
) -> Optional[float]:
    idx: Dict[Tuple[int, ...], float] = lut.get('index', {})
    pts: List[Tuple[int, ...]] = lut.get('points', [])
    path = str(lut.get('path', ''))
    key = tuple(int(max(1, int(x))) for x in dims)
    if key in idx:
        us = float(idx[key])
        _log_lut_hit_once(tag, key, us, path)
        return float(us)
    if pts:
        _warn_missing_once(tag, key, path)
        distance_fn = None
        if str(tag).upper() == 'MMAD':
            cube = int(lut.get('cube', 16))
            residue_weight = float(lut.get('residue_weight', 0.05))
            distance_fn = lambda qq, pp: _mmad_distance(qq, pp, cube=cube, residue_weight=residue_weight)
        us, neigh = _idw_interpolate_with_neighbors(
            pts,
            key,
            k_neigh=int(lut.get('k_neigh', 8)),
            power=float(lut.get('power', 1.0)),
            distance_fn=distance_fn,
        )
        if us is not None:
            _log_lut_interp_once(tag, key, float(us), path, neigh)
        return us
    return None


# --------------------------------------------------- MMAD LUT ---------------------------------------------------------


@lru_cache(maxsize=1)
def _load_mmad_lut() -> Optional[Dict[str, Any]]:
    for p in _candidate_mmad_lut_paths():
        try:
            if (not p) or (not os.path.isfile(p)):
                continue

            obj = _load_table_any(p)
            points: List[Tuple[int, int, int, float]] = []

            # common wrappers
            if isinstance(obj, dict):
                for k in ('entries', 'data', 'table', 'lut'):
                    if k in obj:
                        obj = obj.get(k)
                        break

            if isinstance(obj, list):
                # list of dict rows OR list rows like [M, N, K, us]
                for row in obj:
                    if isinstance(row, dict):
                        m = _coerce_int(_first_present(row, ('M', 'm', '0')))
                        n = _coerce_int(_first_present(row, ('N', 'n', '1')))
                        k = _coerce_int(_first_present(row, ('K', 'k', '2')))
                        us = _coerce_float(_first_present(row, ('total_time_us', 'total_us', 'us', 'lat_us', 'latency_us', 'time_us', '3')))
                        if None not in (m, n, k, us):
                            points.append((int(m), int(n), int(k), float(us)))
                    elif isinstance(row, (list, tuple)) and len(row) >= 4:
                        m = _coerce_int(row[0]); n = _coerce_int(row[1]); k = _coerce_int(row[2]); us = _coerce_float(row[3])
                        if None not in (m, n, k, us):
                            points.append((int(m), int(n), int(k), float(us)))

            elif isinstance(obj, dict):
                # nested dict {M:{N:{K:us}}}
                points = _flatten_nested_numeric_dict_3d(obj)

                if not points:
                    # dict mapping composite-key -> value (e.g., "M,N,K" -> us)
                    for kk, vv in obj.items():
                        if isinstance(vv, dict):
                            continue
                        us = _coerce_float(vv)
                        if us is None:
                            continue
                        nums = re.findall(r"-?\d+", str(kk))
                        if len(nums) >= 3:
                            m = _coerce_int(nums[0]); n = _coerce_int(nums[1]); k = _coerce_int(nums[2])
                            if None not in (m, n, k):
                                points.append((int(m), int(n), int(k), float(us)))

                if not points:
                    # dict of rows keyed by composite string (value is a dict row)
                    for _, row in obj.items():
                        if not isinstance(row, dict):
                            continue
                        m = _coerce_int(_first_present(row, ('M', 'm')))
                        n = _coerce_int(_first_present(row, ('N', 'n')))
                        k = _coerce_int(_first_present(row, ('K', 'k')))
                        us = _coerce_float(_first_present(row, ('total_time_us', 'total_us', 'us', 'lat_us', 'latency_us', 'time_us')))
                        if None not in (m, n, k, us):
                            points.append((int(m), int(n), int(k), float(us)))

            if not points:
                raise ValueError('no valid MMAD LUT points found')

            lut = {
                'path': str(p),
                'points': points,
                'index': _build_index(points),
                # Keep MMAD interpolation local and tile-aware for Ascend Cube.
                'cube': 16,
                'residue_weight': 0.05,
                'k_neigh': 2,
                'power': 1.0,
            }
            logger.info(f"[MMAD-LUT] Loaded {len(points)} points from '{p}'")
            return lut
        except Exception as e:
            logger.debug(f"[MMAD-LUT] Failed to load '{p}': {e}")
    return None


# --------------------------------------------------- Softmax LUT

@lru_cache(maxsize=1)
def _load_softmax_lut() -> Optional[Dict[str, Any]]:
    for p in _candidate_softmax_lut_paths():
        try:
            if (not p) or (not os.path.isfile(p)):
                continue

            obj = _load_table_any(p)
            points: List[Tuple[int, int, float]] = []

            if isinstance(obj, dict):
                for k in ('entries', 'data', 'table', 'lut'):
                    if k in obj:
                        obj = obj.get(k)
                        break

            if isinstance(obj, list):
                for row in obj:
                    if isinstance(row, dict):
                        m = _coerce_int(_first_present(row, ('M', 'm', '0')))
                        n = _coerce_int(_first_present(row, ('N', 'n', 'K', 'k', '1')))
                        us = _coerce_float(_first_present(row, ('total_time_us', 'total_us', 'us', 'lat_us', 'latency_us', 'time_us', '2')))
                        if None not in (m, n, us):
                            points.append((int(m), int(n), float(us)))
                    elif isinstance(row, (list, tuple)) and len(row) >= 3:
                        m = _coerce_int(row[0]); n = _coerce_int(row[1]); us = _coerce_float(row[2])
                        if None not in (m, n, us):
                            points.append((int(m), int(n), float(us)))

            elif isinstance(obj, dict):
                # nested dict {M:{N:us}}
                for mk, mv in obj.items():
                    m = _coerce_int(mk)
                    if m is None or not isinstance(mv, dict):
                        continue
                    for nk, vv in mv.items():
                        n = _coerce_int(nk)
                        us = _coerce_float(vv)
                        if None not in (m, n, us):
                            points.append((int(m), int(n), float(us)))

            if not points:
                raise ValueError('no valid Softmax LUT points found')

            lut = {
                'path': str(p),
                'points': points,
                'index': _build_index(points),
                'k_neigh': 8,
                'power': 1.0,
            }
            logger.info(f"[SOFTMAX-LUT] Loaded {len(points)} points from '{p}'")
            return lut
        except Exception as e:
            logger.debug(f"[SOFTMAX-LUT] Failed to load '{p}': {e}")
    return None

# --------------------------------------------------- GeLU LUT ---------------------------------------------------------
@lru_cache(maxsize=1)
def _load_gelu_lut() -> Optional[Dict[str, Any]]:
    for p in _candidate_gelu_lut_paths():
        try:
            if (not p) or (not os.path.isfile(p)):
                continue

            obj = _load_table_any(p)
            points: List[Tuple[int, float]] = []

            if isinstance(obj, dict):
                for k in ('entries', 'data', 'table', 'lut'):
                    if k in obj:
                        obj = obj.get(k)
                        break

            if isinstance(obj, list):
                for row in obj:
                    if isinstance(row, dict):
                        l = _coerce_int(_first_present(row, ('len', 'length', 'L', 'M', 'm', '0')))
                        us = _coerce_float(_first_present(row, ('total_time_us', 'total_us', 'us', 'lat_us', 'latency_us', 'time_us', '1')))
                        if None not in (l, us):
                            points.append((int(l), float(us)))
                    elif isinstance(row, (list, tuple)) and len(row) >= 2:
                        l = _coerce_int(row[0]); us = _coerce_float(row[1])
                        if None not in (l, us):
                            points.append((int(l), float(us)))

            elif isinstance(obj, dict):
                # dict {len:us}
                for lk, vv in obj.items():
                    l = _coerce_int(lk)
                    us = _coerce_float(vv)
                    if None not in (l, us):
                        points.append((int(l), float(us)))

            if not points:
                raise ValueError('no valid GeLU LUT points found')

            lut = {
                'path': str(p),
                'points': points,
                'index': _build_index(points),
                'k_neigh': 4,
                'power': 1.0,
            }
            logger.info(f"[GELU-LUT] Loaded {len(points)} points from '{p}'")
            return lut
        except Exception as e:
            logger.debug(f"[GELU-LUT] Failed to load '{p}': {e}")
    return None
# --------------------------------------------------- LayerNorm LUT ----------------------------------------------------
@lru_cache(maxsize=1)
def _load_layernorm_lut() -> Optional[Dict[str, Any]]:
    for p in _candidate_layernorm_lut_paths():
        try:
            if (not p) or (not os.path.isfile(p)):
                continue

            obj = _load_table_any(p)
            points: List[Tuple[int, int, float]] = []

            if isinstance(obj, dict):
                for k in ('entries', 'data', 'table', 'lut'):
                    if k in obj:
                        obj = obj.get(k)
                        break

            if isinstance(obj, list):
                for row in obj:
                    if isinstance(row, dict):
                        # Support both:
                        #   (rows, width) table: rows=M, width=N
                        #   (batch, seq, width) table (common for RMSNorm): M=batch, N=seq, K=width
                        mM = _coerce_int(_first_present(row, ('M', 'm', '0')))
                        nN = _coerce_int(_first_present(row, ('N', 'n', '1')))
                        kK = _coerce_int(_first_present(row, ('K', 'k', '2')))

                        if None not in (mM, nN, kK) and int(mM) > 0 and int(nN) > 0 and int(kK) > 0:
                            rows = int(mM) * int(nN)
                            width = int(kK)
                        else:
                            rows = _coerce_int(_first_present(row, ('rows', 'row', 'M', 'm', '0')))
                            width = _coerce_int(_first_present(row, ('width', 'dim', 'D', 'K', 'k', 'N', 'n', '1')))

                        us = _coerce_float(_first_present(row, ('total_time_us', 'total_us', 'us', 'lat_us', 'latency_us', 'time_us', '2', '3')))
                        if None not in (rows, width, us):
                            points.append((int(rows), int(width), float(us)))

                    elif isinstance(row, (list, tuple)) and len(row) >= 3:
                        m = _coerce_int(row[0]); n = _coerce_int(row[1]); us = _coerce_float(row[2])
                        if None not in (m, n, us):
                            points.append((int(m), int(n), float(us)))

            elif isinstance(obj, dict):
                # nested dict {rows:{width:us}}
                for mk, mv in obj.items():
                    m = _coerce_int(mk)
                    if m is None or not isinstance(mv, dict):
                        continue
                    for nk, vv in mv.items():
                        n = _coerce_int(nk)
                        us = _coerce_float(vv)
                        if None not in (m, n, us):
                            points.append((int(m), int(n), float(us)))

            if not points:
                raise ValueError('no valid Norm/LN/RMSNorm LUT points found')

            lut = {
                'path': str(p),
                'points': points,
                'index': _build_index(points),
                'k_neigh': 8,
                'power': 1.0,
            }
            logger.info(f"[NORM-LUT] Loaded {len(points)} points from '{p}'")
            return lut
        except Exception as e:
            logger.debug(f"[NORM-LUT] Failed to load '{p}': {e}")
    return None

# --------------------------------------------------- Legacy regression helpers ----------------------------------------

def _ceil_div(a: int, b: int) -> int:
    return (a + b - 1) // b


def _compute_feature_vector(M: int, N: int, K: int, block_size: int, feature_names: list) -> Tuple[list, list]:
    MB = _ceil_div(M, block_size)
    NB = _ceil_div(N, block_size)
    KB = _ceil_div(K, block_size)
    base = {
        'MB': MB, 'NB': NB, 'KB': KB,
        'tiles': MB * NB * KB,
        'mn': MB * NB,
        'sum_b': MB + NB + KB,
        'M': M, 'N': N, 'K': K,
    }
    feats = [float(base[name]) for name in feature_names]
    return (feats, [MB, NB, KB])



def _map_op_to_mmad_dims(
    op: str,
    dim: int,
    n_heads: int,
    n_kv_heads: int,
    ffn_dim: int,
    seqlen: int,
    batch: int = 1,
    q_len: Optional[int] = None,
    kv_len: Optional[int] = None,
    phase: Optional[str] = None,
) -> Optional[Tuple[int, int, int, int]]:

    if not op:
        logger.debug(str(f"[MAP] op is None/empty, returning None"))
        return None

    op = str(op).strip().lower()

    try:
        dim_i = int(dim)
        n_heads_i = int(max(1, n_heads))
        n_kv_i = int(max(1, n_kv_heads))
        ffn_i = int(ffn_dim)
        seq_i = int(max(1, seqlen))
        batch_i = int(max(1, batch))
    except Exception:
        logger.debug(str(f"[MAP] invalid dims for op='{op}'"))
        return None

    ph = str(phase or "").strip().lower()

    # Infer q_len / kv_len with backward-compatible defaults.
    if q_len is None:
        q_i = 1 if ph == "decode" else seq_i
    else:
        q_i = int(max(1, q_len))

    if kv_len is None:
        kv_i = seq_i
    else:
        kv_i = int(max(1, kv_len))

    head_dim = int(max(1, dim_i // n_heads_i))
    kv_dim = int(max(1, n_kv_i * head_dim))
    ffn_out = int(max(1, ffn_i if ffn_i > 0 else 4 * dim_i))

    # Fold repeats into M to match the LUT benchmark behavior.
    M_q = int(batch_i * q_i)  # GEMM: [M_q, K] x [K, N]

    if op == "q_proj":
        return (M_q, dim_i, dim_i, 1)
    if op in ("k_proj", "v_proj"):
        return (M_q, kv_dim, dim_i, 1)
    if op == "wo_proj":
        return (M_q, dim_i, dim_i, 1)
    if op in ("ffn_up", "ffn_gate"):
        return (M_q, ffn_out, dim_i, 1)
    if op == "ffn_down":
        return (M_q, dim_i, ffn_out, 1)
    if op == "score":
        M = int(batch_i * n_heads_i * q_i)
        return (M, kv_i, head_dim, 1)
    if op == "output":
        M = int(batch_i * n_heads_i * q_i)
        return (M, head_dim, kv_i, 1)

    logger.debug(str(f"[MAP-MMAD] No match for op='{op}'"))
    return None

# --------------------------------------------------- Public predictor APIs (LUT) --------------------------------------

def _predict_mmad_latency_us_from_lut(M: int, N: int, K: int) -> Optional[float]:
    lut = _load_mmad_lut()
    if not lut:
        return None
    return _lut_query(tag='MMAD', lut=lut, dims=(int(M), int(N), int(K)))


def _predict_softmax_latency_us_from_lut(M: int, K: int, *, phase: Optional[str]=None, causal: Optional[bool]=None) -> Optional[float]:
    """Softmax LUT query (phase/causal are accepted for signature compatibility, but LUT is shape-only)."""
    lut = _load_softmax_lut()
    if not lut:
        return None
    return _lut_query(tag='SOFTMAX', lut=lut, dims=(int(M), int(K)))


def _predict_gelu_latency_us_from_lut(data_len: int) -> Optional[float]:
    lut = _load_gelu_lut()
    if not lut:
        return None
    return _lut_query(tag='GELU', lut=lut, dims=(int(data_len),))


def _predict_layernorm_latency_us_from_lut(rows: int, width: int) -> Optional[float]:
    """Norm LUT query: treat RMSNorm LUT as generic norm (LN/RMSNorm/GroupNorm...)."""
    lut = _load_layernorm_lut()
    if not lut:
        return None
    return _lut_query(tag='NORM', lut=lut, dims=(int(rows), int(width)))

# Canonical names (recommended)
_predict_mmad_latency_us = _predict_mmad_latency_us_from_lut
_predict_softmax_latency_us = _predict_softmax_latency_us_from_lut
_predict_gelu_latency_us = _predict_gelu_latency_us_from_lut
_predict_layernorm_latency_us = _predict_layernorm_latency_us_from_lut
_predict_norm_latency_us = _predict_layernorm_latency_us_from_lut

